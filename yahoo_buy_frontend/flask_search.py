import sys, os
sys.path.append('.')

from flask import Flask, render_template
#  引入form類別
from search_form import SearchForm
from flask import request

app = Flask(__name__)

def _get_db_info():
    import os
    with open(str(os.getcwd()) + '/yahoo_buy_db.json', encoding='utf-8') as f:
        import json
        db_data = json.load(f)

    return db_data

def _select_update_from_db(sql, parameters=''):
    from time import sleep

    db_data = _get_db_info()
    conn = None
    cur = None
    retry_limit = 5
    count = 0
    while count < retry_limit:
        try:
            import pymysql
            conn = pymysql.connect(host=db_data['to_db']['host'],
                                   port=db_data['to_db']['port'],
                                   user=db_data['to_db']['user'],
                                   passwd=db_data['to_db']['pass'],
                                   db=db_data['to_db']['name'],
                                   charset=db_data['to_db']['charset'])
            cur = conn.cursor()
            if parameters:
                cur.execute(sql, parameters)
            else:
                cur.execute(sql)

            results = cur.fetchall()
            # for row in results:
            #     print(row)

            conn.commit()
            break
        except:
            import traceback
            traceback.print_exc()

            print('Fetching database fails. Trying...')
            count += 1
            sleep(1)

    cur.close()
    conn.close()

    if count >= retry_limit:
        return 'fetch_db_error'
    else:
        return results

def _get_data_from_db(query_text, input_sort=''):
    print(input_sort)
    if input_sort!='all':
        sql = "SELECT category,item_title,item_price,item_info,item_url,SEC_TO_TIME(UNIX_TIMESTAMP(check_time)-UNIX_TIMESTAMP(created_time)) AS staying_period FROM item_information WHERE " + input_sort + " LIKE '%" + query_text + "%' ORDER BY (UNIX_TIMESTAMP(check_time)-UNIX_TIMESTAMP(created_time)) DESC"
    else:
        sql = "SELECT category,item_title,item_price,item_info,item_url,SEC_TO_TIME(UNIX_TIMESTAMP(check_time)-UNIX_TIMESTAMP(created_time)) AS staying_period FROM item_information WHERE CONCAT(category, item_title, item_info) LIKE '%" + query_text + "%' ORDER BY (UNIX_TIMESTAMP(check_time)-UNIX_TIMESTAMP(created_time)) DESC"
    results = _select_update_from_db(sql)
    return results

@app.route('/output_csv', methods=['GET', 'POST'])
def _output_csv():
    query_text = request.args.get('query_text')
    input_sort = request.args.get('input_sort')
    results = _get_data_from_db(query_text, input_sort)
    if results=='fetch_db_error':
        print('Fetching database fails. Please check the log...')
        return ''

    # write header to csv file
    import datetime
    import csv

    now = datetime.datetime.now()
    filename = 'yahoo_buy_search_result_' + now.strftime("%Y-%m-%d") + '.csv'
    file = open('static/' + filename, 'w', newline='', encoding='utf8')
    csvCursor = csv.writer(file)
    csvHeader = ['category', 'item_title', 'item_price', 'item_info', 'item_url', 'staying_period']
    csvCursor.writerow(csvHeader)
    csvCursor.writerow(['-----', '-----', '-----', '-----', '-----', '-----'])

    for one_result in results:
        item_info_list = ['‧' + each_line for each_line in one_result[3].split(';;;')]

        control_characters = dict.fromkeys(range(32))
        remove_control = lambda x: x.translate(control_characters)
        csvCursor.writerow([remove_control(one_result[0]), remove_control(one_result[1]), remove_control(one_result[2]), remove_control("\r\n".join(item_info_list)), one_result[4], '\"' + str(one_result[5]) + '\"'])

    file.close()
    # ___ Output to a CSV file

    from flask import send_from_directory
    return send_from_directory('static', filename)

@app.route('/output_excel', methods=['GET', 'POST'])
def _output_excel():
    query_text = request.args.get('query_text')
    input_sort = request.args.get('input_sort')
    results = _get_data_from_db(query_text, input_sort)
    if results=='fetch_db_error':
        print('Fetching database fails. Please check the log...')
        return ''

    # *** Output to a Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, Color
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = 'Yahoo Buy Search Results'  # 設置worksheet的標題

    # write header
    excelHeader = ['category', 'item_title', 'item_price', 'item_info', 'staying_period']
    ws.append(excelHeader)
    ws.append(['-----', '-----', '-----', '-----', '-----'])

    ft = Font()
    ft.underline = 'single'  # add single underline
    ft.color = Color(rgb='000000FF')  # add blue color
    for one_result in results:
        item_info_list = ['‧' + each_line for each_line in one_result[3].split(';;;')]

        control_characters = dict.fromkeys(range(32))
        remove_control = lambda x: x.translate(control_characters)
        ws.append([remove_control(one_result[0]), remove_control(one_result[1]), remove_control(one_result[2]), remove_control("\r\n".join(item_info_list)), one_result[5]])

        ws['A' + str(ws._current_row)] = '=HYPERLINK("{}", "{}")'.format(one_result[4], one_result[0])
        ws['A' + str(ws._current_row)].font = ft

    import datetime
    filename = 'yahoo_buy_search_result_' + datetime.datetime.now().strftime("%Y-%m-%d") + '.xlsx'
    wb.save('static/' + filename)
    # ___ Output to a Excel file

    from flask import send_from_directory
    return send_from_directory('static', filename)


@app.route('/', methods=['GET', 'POST'])
def yahoo_buy_search():
    form = SearchForm()
    #  flask_wtf類中提供判斷是否表單提交過來的method，不需要自行利用request.method來做判斷
    if form.validate_on_submit():
        return render_template('yahoo_buy_search.html', form=form, result_div=_get_data_from_db(form.input_question.data, form.input_sort.data))
    #  如果不是提交過來的表單，就是GET，這時候就回傳yahoo_buy_search.html網頁
    return render_template('yahoo_buy_search.html', form=form)

if __name__ == '__main__':
    app.debug = True
    app.config['SECRET_KEY'] = 'your key values'
    app.run(host='0.0.0.0')